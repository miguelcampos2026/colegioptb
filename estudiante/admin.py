# Register your models here.
from django.contrib import admin
from django.utils.html import format_html
from .models import Estudiante, Matricula, Colegiatura, PendientePago, Pago, Reporte
from django.urls import path
from django.shortcuts import render
from django.utils.timezone import now
from .models import PendientePago
from django.db.models import Sum
from datetime import datetime
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template
# utilidades para PDF y Excel (pegar cerca de los imports)
import os
import io
from django.conf import settings
from django.contrib.staticfiles import finders
from pathlib import Path

def link_callback(uri, rel):
    """
    Resuelve una URI de static/media a una ruta absoluta en disco
    para que xhtml2pdf pueda abrir la imagen.
    """
    # URLs remotas
    if uri.startswith("http://") or uri.startswith("https://"):
        return uri

    # Si ya es file:/// limpiarla
    if uri.startswith("file:///"):
        path = uri.replace("file:///", "")
        if os.path.exists(path):
            return path

    # Candidate sin prefijo STATIC_URL
    candidate = uri
    if uri.startswith(settings.STATIC_URL):
        candidate = uri[len(settings.STATIC_URL):].lstrip("/")

    # Intentar con finders (lo más fiable)
    found = finders.find(candidate)
    if found:
        return found

    # Intentar STATIC_ROOT
    if getattr(settings, "STATIC_ROOT", None):
        path = os.path.join(settings.STATIC_ROOT, candidate)
        if os.path.exists(path):
            return path

    # Intentar STATICFILES_DIRS
    for static_dir in getattr(settings, "STATICFILES_DIRS", []):
        path = os.path.join(static_dir, candidate)
        if os.path.exists(path):
            return path

    # Intentar relativo a BASE_DIR
    path = os.path.join(settings.BASE_DIR, candidate)
    if os.path.exists(path):
        return path

    # Si no se encuentra, lanzar excepción para facilitar debug
    raise FileNotFoundError(f"No se pudo localizar el archivo estático: {uri}")

def render_pdf_response(template_name, context, filename="reporte.pdf"):
    """
    Renderiza template a PDF y devuelve HttpResponse listo.
    """
    template = get_template(template_name)
    html = template.render(context)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename={filename}"
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisa_status.err:
        return HttpResponse("Error al generar el PDF", status=500)
    return response

@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    list_display = ("codigo", "nombres", "apellidos", "nivelogrado", "seccion", "activo")
    search_fields = ("codigo", "nombres", "apellidos")
    list_filter = ("nivelogrado", "activo")
    def has_delete_permission(self, request, obj=None):
        return False

@admin.register(Matricula)
class MatriculaAdmin(admin.ModelAdmin):
    list_display = ("estudiante", "anio_lectivo", "fecha_matricula", "total_matricula", "colegiatura_mensual")
    search_fields = ("estudiante__nombres", "estudiante__apellidos")
    list_filter = ("anio_lectivo",)
    def has_delete_permission(self, request, obj=None):
        return False

@admin.register(Colegiatura)
class ColegiaturaAdmin(admin.ModelAdmin):
    list_display = ("matricula", "mes", "monto", "pagada", "fecha_pago")
    list_filter = ("mes", "pagada")
    search_fields = ("matricula__estudiante__nombres", "matricula__estudiante__apellidos")

@admin.register(PendientePago)
class PendientePagoAdmin(admin.ModelAdmin):
    list_display = ("estudiante", "concepto", "monto", "recargo", "pagado", "fecha_vencimiento", "fecha_pago")
    list_filter = ("concepto", "pagado")
    search_fields = ("estudiante__nombres", "estudiante__apellidos", "concepto")

@admin.register(Pago)
class PagoAdmin(admin.ModelAdmin):
    list_display = (
        "recibo_numero",
        "estudiante",
        "fecha_pago",
        "responsable",
        "total_pago",
        "ver_recibo") # Aquí agregamos la columna personalizada
    filter_horizontal = ("pendientes",)  # selector de múltiples pendientes

    def ver_recibo(self, obj):
        # Genera un enlace al recibo en la vista personalizada
        return format_html('<a href="/estudiante/recibo/{}/" target="_blank">Imprimir</a>', obj.id)
        
        ver_recibo.short_description = "Recibo"

# --- Clase ReportesAdmin completa ---
@admin.register(Reporte)
class ReportesAdmin(admin.ModelAdmin):
    change_list_template = "reportes/reportes_menu.html"

    def get_queryset(self, request):
        return Reporte.objects.none()

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("", self.admin_site.admin_view(self.reportes_menu), name="admin_reportes_menu"),
            path("form/<slug:slug>/", self.admin_site.admin_view(self.reporte_form), name="admin_reporte_form"),
            path("form/<slug:slug>/pdf/", self.admin_site.admin_view(self.reporte_export_pdf), name="admin_reporte_export_pdf"),
            path("form/<slug:slug>/excel/", self.admin_site.admin_view(self.reporte_export_excel), name="admin_reporte_export_excel"),
            # Rutas antiguas compatibles
            path("pendientes/", self.admin_site.admin_view(self.reporte_pendientes), name="admin_reporte_pendientes"),
            path("morosidad/", self.admin_site.admin_view(self.reporte_morosidad), name="admin_reporte_morosidad"),
            path("pagos-mensuales/", self.admin_site.admin_view(self.reporte_pagos_mensuales), name="admin_reporte_pagos_mensuales"),
            path("balance/", self.admin_site.admin_view(self.reporte_balance_general), name="admin_reporte_balance_general"),
            path("morosidad/pdf/", self.admin_site.admin_view(self.reporte_morosidad_pdf), name="admin_reporte_morosidad_pdf"),
        ]
        return custom_urls + urls

    # ---------- Menú y formulario genérico ----------
    def reportes_menu(self, request):
        return render(request, "reportes/reportes_menu.html")

    def reporte_form(self, request, slug):
        """
        Vista genérica que muestra el formulario de parámetros y la vista previa.
        Si se envía ?generar=1 se renderiza la vista previa HTML del reporte.
        """
        from datetime import datetime, date

        titulo_map = {
            "morosidad": "Reporte de Morosidad",
            "pendientes": "Reporte de Pagos Pendientes",
            "pagos-mensuales": "Pagos Realizados por Mes",
            "balance": "Balance General de Ingresos",
        }
        titulo = titulo_map.get(slug, "Reporte")

        context_base = {
            "slug": slug,
            "titulo": titulo,
            "logo_path": settings.STATIC_URL + "img/logo_colegio.png",
            "request": request,
        }

        # Si se solicita generar vista previa
        if request.GET.get("generar"):
            # Reusar lógica por slug para construir contexto específico
            if slug == "morosidad":
                ctx = self._get_morosidad_context(request)
                context_base.update(ctx)
                return render(request, "reportes/reporte_morosidad.html", context_base)
            if slug == "pendientes":
                ctx = self._get_pendientes_context(request)
                context_base.update(ctx)
                return render(request, "reportes/reporte_pendientes.html", context_base)
            if slug == "pagos-mensuales":
                ctx = self._get_pagos_mensuales_context(request)
                context_base.update(ctx)
                return render(request, "reportes/reporte_pagos_mensuales.html", context_base)
            if slug == "balance":
                ctx = self._get_balance_context(request)
                context_base.update(ctx)
                return render(request, "reportes/reporte_balance.html", context_base)

            # fallback
            return render(request, "reportes/reporte_form.html", context_base)

        # Si no se generó, mostrar formulario con botones de exportación
        return render(request, "reportes/reporte_form.html", context_base)

    # ---------- Exportaciones genéricas ----------
    def reporte_export_pdf(self, request, slug):
        # Construir contexto según slug y llamar a render_pdf_response
        if slug == "morosidad":
            ctx = self._get_morosidad_context(request)
            template = "reportes/pdf_morosidad.html"
            filename = "reporte_morosidad.pdf"
        elif slug == "pendientes":
            ctx = self._get_pendientes_context(request)
            template = "reportes/pdf_pendientes.html"
            filename = "reporte_pendientes.pdf"
        elif slug == "pagos-mensuales":
            ctx = self._get_pagos_mensuales_context(request)
            template = "reportes/pdf_pagos_mensuales.html"
            filename = "reporte_pagos_mensuales.pdf"
        elif slug == "balance":
            ctx = self._get_balance_context(request)
            template = "reportes/pdf_balance_general.html"
            filename = "reporte_balance_general.pdf"
        else:
            return HttpResponse("Reporte no encontrado", status=404)

        # Añadir datos comunes
        ctx.update({
            "titulo_reporte": ctx.get("titulo_reporte", ""),
            "fecha_emision": now().date(),
            "logo_path": settings.STATIC_URL + "img/logo_colegio.png",
        })
        return render_pdf_response(template, ctx, filename=filename)

    def reporte_export_excel(self, request, slug):
        # Exportar a Excel usando openpyxl
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("Instale openpyxl para exportar a Excel", status=500)

        if slug == "morosidad":
            ctx = self._get_morosidad_context(request)
            filename = "reporte_morosidad.xlsx"
            # Construir workbook simple
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Morosidad"
            headers = ["Estudiante", "Meses Pendientes", "Total Vencido", "Detalle"]
            for col, h in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = h
            row = 2
            for e in ctx["estudiantes"]:
                detalle = "; ".join([f"{p.concepto} - vence {p.fecha_vencimiento} ({p.monto})" for p in e["pendientes"]])
                ws[f"A{row}"] = f"{e['nombres']} {e['apellidos']}"
                ws[f"B{row}"] = e["meses_pendientes"]
                ws[f"C{row}"] = float(e["total_vencido"])
                ws[f"D{row}"] = detalle
                row += 1

        elif slug == "pendientes":
            ctx = self._get_pendientes_context(request)
            filename = "reporte_pendientes.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pendientes"
            headers = ["Estudiante", "Concepto", "Monto", "Fecha Vencimiento"]
            for col, h in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = h
            row = 2
            for e in ctx["estudiantes"]:
                for p in e["pendientes"]:
                    ws[f"A{row}"] = f"{e['nombres']} {e['apellidos']}"
                    ws[f"B{row}"] = p.concepto
                    ws[f"C{row}"] = float(p.monto)
                    ws[f"D{row}"] = p.fecha_vencimiento.strftime("%Y-%m-%d")
                    row += 1

        elif slug == "pagos-mensuales":
            ctx = self._get_pagos_mensuales_context(request)
            filename = "reporte_pagos_mensuales.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pagos Mensuales"
            headers = ["Recibo", "Estudiante", "Fecha Pago", "Responsable", "Total"]
            for col, h in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = h
            row = 2
            for p in ctx["pagos"]:
                ws[f"A{row}"] = p.recibo_numero
                ws[f"B{row}"] = f"{p.estudiante.nombres} {p.estudiante.apellidos}"
                ws[f"C{row}"] = p.fecha_pago.strftime("%Y-%m-%d")
                ws[f"D{row}"] = p.responsable
                ws[f"E{row}"] = float(p.total_pago)
                row += 1

        elif slug == "balance":
            ctx = self._get_balance_context(request)
            filename = "reporte_balance_general.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Balance"
            headers = ["Recibo", "Estudiante", "Total Pendientes"]
            for col, h in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = h
            row = 2
            for pago, total in ctx["pagos_totales"]:
                ws[f"A{row}"] = pago.recibo_numero
                ws[f"B{row}"] = f"{pago.estudiante.nombres} {pago.estudiante.apellidos}"
                ws[f"C{row}"] = float(total)
                row += 1

        else:
            return HttpResponse("Reporte no encontrado", status=404)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        response = HttpResponse(stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    # ---------- Métodos auxiliares para construir contextos por reporte ----------
    def _get_morosidad_context(self, request):
        """
        Devuelve contexto con 'estudiantes' (lista de dicts) y 'fecha_consulta'
        """
        from datetime import datetime, date
        fecha_str = request.GET.get("fecha")
        if fecha_str:
            formatos = ("%Y-%m-%d","%b. %d, %Y","%b %d, %Y","%d/%m/%Y","%m/%d/%Y")
            fecha_consulta = None
            for fmt in formatos:
                try:
                    fecha_consulta = datetime.strptime(fecha_str, fmt).date()
                    break
                except ValueError:
                    continue
            if fecha_consulta is None:
                fecha_consulta = date.today()
        else:
            fecha_consulta = date.today()

        pendientes = PendientePago.objects.filter(
            pagado=False,
            concepto__icontains="Colegiatura",
            fecha_vencimiento__lt=fecha_consulta
        ).select_related("estudiante")

        estudiantes = {}
        for p in pendientes:
            if p.estudiante.id not in estudiantes:
                estudiantes[p.estudiante.id] = {
                    "nombres": p.estudiante.nombres,
                    "apellidos": p.estudiante.apellidos,
                    "pendientes": [],
                    "total_vencido": 0,
                    "meses_pendientes": 0,
                }
            estudiantes[p.estudiante.id]["pendientes"].append(p)
            estudiantes[p.estudiante.id]["total_vencido"] += p.monto
            estudiantes[p.estudiante.id]["meses_pendientes"] += 1

        return {
            "estudiantes": list(estudiantes.values()),
            "fecha_consulta": fecha_consulta,
            "titulo_reporte": "Reporte de Morosidad",
        }

    def _get_pendientes_context(self, request):
        from datetime import datetime, date
        fecha_str = request.GET.get("fecha")
        if fecha_str:
            formatos = ("%Y-%m-%d","%b. %d, %Y","%b %d, %Y","%d/%m/%Y","%m/%d/%Y")
            fecha_consulta = None
            for fmt in formatos:
                try:
                    fecha_consulta = datetime.strptime(fecha_str, fmt).date()
                    break
                except ValueError:
                    continue
            if fecha_consulta is None:
                fecha_consulta = date.today()
        else:
            fecha_consulta = date.today()

        pendientes = PendientePago.objects.filter(
            pagado=False,
            fecha_vencimiento__lte=fecha_consulta
        ).select_related("estudiante")

        estudiantes = {}
        for p in pendientes:
            if p.estudiante.id not in estudiantes:
                estudiantes[p.estudiante.id] = {
                    "nombres": p.estudiante.nombres,
                    "apellidos": p.estudiante.apellidos,
                    "pendientes": []
                }
            estudiantes[p.estudiante.id]["pendientes"].append(p)

        return {
            "estudiantes": list(estudiantes.values()),
            "fecha_consulta": fecha_consulta,
            "titulo_reporte": "Reporte de Pagos Pendientes",
        }

    def _get_pagos_mensuales_context(self, request):
        hoy = now().date()
        mes = request.GET.get("mes")
        anio = request.GET.get("anio")
        if mes and anio:
            try:
                mes = int(mes); anio = int(anio)
            except ValueError:
                mes = hoy.month; anio = hoy.year
        else:
            mes = hoy.month; anio = hoy.year

        pagos = Pago.objects.filter(fecha_pago__month=mes, fecha_pago__year=anio).select_related("estudiante")
        return {
            "pagos": pagos,
            "mes": mes,
            "anio": anio,
            "titulo_reporte": f"Pagos Mensuales {mes}/{anio}",
        }

    def _get_balance_context(self, request):
        pagos = Pago.objects.all().prefetch_related("pendientes")
        pagos_totales = []
        for pago in pagos:
            total = sum(p.monto for p in pago.pendientes.all())
            pagos_totales.append((pago, total))
        total_ingresos = sum(total for _, total in pagos_totales)
        return {
            "pagos_totales": pagos_totales,
            "balance": {"total_ingresos": total_ingresos},
            "titulo_reporte": "Balance General",
        }

    # ---------- Métodos legacy para compatibilidad con tu menú actual ----------
    # (mantengo tus vistas originales para que no rompan enlaces existentes)
    def reporte_pendientes(self, request):
        ctx = self._get_pendientes_context(request)
        return render(request, "reportes/reporte_pendientes.html", ctx)

    def reporte_morosidad(self, request):
        ctx = self._get_morosidad_context(request)
        return render(request, "reportes/reporte_morosidad.html", ctx)

    def reporte_morosidad_pdf(self, request):
        ctx = self._get_morosidad_context(request)
        ctx.update({"logo_path": settings.STATIC_URL + "img/logo_colegio.png", "fecha_emision": now().date()})
        return render_pdf_response("reportes/pdf_morosidad.html", ctx, filename="reporte_morosidad.pdf")

    def reporte_pagos_mensuales(self, request):
        ctx = self._get_pagos_mensuales_context(request)
        return render(request, "reportes/reporte_pagos_mensuales.html", ctx)

    def reporte_balance_general(self, request):
        ctx = self._get_balance_context(request)
        return render(request, "reportes/reporte_balance.html", ctx)
